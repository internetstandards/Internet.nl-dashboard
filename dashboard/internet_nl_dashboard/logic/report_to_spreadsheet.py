import itertools
import logging
from string import ascii_uppercase
from tempfile import NamedTemporaryFile
from typing import Any, Dict, List, Union

import pyexcel as p
from django.utils.text import slugify
from openpyxl import load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, PatternFill

from dashboard.internet_nl_dashboard.logic import (MAIL_AUTH_CATEGORY, MAIL_AUTH_FIELDS,
                                                   MAIL_DNSSEC_CATEGORY, MAIL_DNSSEC_FIELDS,
                                                   MAIL_IPV6_CATEGORY, MAIL_IPV6_FIELDS,
                                                   MAIL_LEGACY_FIELDS, MAIL_OVERALL_FIELDS,
                                                   MAIL_TLS_CATEGORY, MAIL_TLS_CERTIFICATE_FIELDS,
                                                   MAIL_TLS_DANE_FIELDS, MAIL_TLS_TLS_FIELDS,
                                                   WEB_APPSECPRIV_CATEGORY, WEB_APPSECPRIV_FIELDS,
                                                   WEB_DNSSEC_CATEGORY, WEB_DNSSEC_FIELDS,
                                                   WEB_IPV6_CATEGORY, WEB_IPV6_FIELDS,
                                                   WEB_LEGACY_FIELDS, WEB_OVERALL_FIELDS,
                                                   WEB_TLS_CATEGORY, WEB_TLS_CERTIFICATE_FIELDS,
                                                   WEB_TLS_DANE_FIELDS, WEB_TLS_HTTP_FIELDS,
                                                   WEB_TLS_TLS_FIELDS)
from dashboard.internet_nl_dashboard.logic.internet_nl_translations import (get_po_as_dictionary_v2,
                                                                            translate_field)
from dashboard.internet_nl_dashboard.models import Account, UrlListReport

log = logging.getLogger(__package__)

# todo: read the preferred language of the user and use the translations matching this user...
po_file_as_dictionary = get_po_as_dictionary_v2('en')

"""
Creates spreadsheets containing report data.

Done: make sure the columns are in a custom order. Columns are defined in scanners. The order of the columns isn't.
Done: make sure you can download the spreadsheet in a single click
Done: add a sane logic to the spreadsheet content
Done: support ods and xlsx.
Done: can we support sum rows at the top for boolean values(?), only for excel. Not for ods, which is too bad.
Done: get column translations. From internet.nl PO files? How do they map to these variables?
Done: write some tests for these methods, once they are more table.
"""

SANE_COLUMN_ORDER = {
    # scanner
    'dns_a_aaaa': {
        'overall': WEB_OVERALL_FIELDS,

        'ipv6': WEB_IPV6_CATEGORY + WEB_IPV6_FIELDS,

        'dnssec': WEB_DNSSEC_CATEGORY + WEB_DNSSEC_FIELDS,

        'tls': WEB_TLS_CATEGORY + WEB_TLS_HTTP_FIELDS + WEB_TLS_TLS_FIELDS + WEB_TLS_CERTIFICATE_FIELDS +
        WEB_TLS_DANE_FIELDS,

        # Added 24th of May 2019
        'appsecpriv': WEB_APPSECPRIV_CATEGORY + WEB_APPSECPRIV_FIELDS,

        'legacy': WEB_LEGACY_FIELDS
    },
    'dns_soa': {
        # any grouping, every group has a empty column between them. The label is not used.
        'overall': MAIL_OVERALL_FIELDS,
        'ipv6': MAIL_IPV6_CATEGORY + MAIL_IPV6_FIELDS,

        'dnssec': MAIL_DNSSEC_CATEGORY + MAIL_DNSSEC_FIELDS,

        'auth': MAIL_AUTH_CATEGORY + MAIL_AUTH_FIELDS,

        # perhaps split these into multiple groups.
        'tls': MAIL_TLS_CATEGORY + MAIL_TLS_TLS_FIELDS + MAIL_TLS_CERTIFICATE_FIELDS + MAIL_TLS_DANE_FIELDS,

        'legacy': MAIL_LEGACY_FIELDS
    },
}


def iter_all_strings():
    # https://stackoverflow.com/questions/29351492/how-to-make-a-continuous-alphabetic-list-python-from-a-z-then-from-aa-ab
    for size in itertools.count(1):
        for product in itertools.product(ascii_uppercase, repeat=size):
            yield "".join(product)


def create_spreadsheet(account: Account, report_id: int):
    # Fails softly, without exceptions if there is nothing yet.
    report = UrlListReport.objects.all().filter(
        urllist__account=account,
        pk=report_id).select_related('urllist').first()

    if not report:
        return None, None

    calculation = report.calculation
    urls = calculation["urls"]

    protocol = 'dns_soa' if report.report_type == 'mail' else 'dns_a_aaaa'

    # results is a matrix / 2-d array / array with arrays.
    data: List[List[Any]] = []
    # done: functions are not supported in ods, which makes the output look terrible... Should we omit this?
    # omitted, as it gives inconsistent results. Kept code to re-use it if there is a workaround for ods / or we
    # know this is an ods file.
    # data += [formula_row(protocol=protocol, function="=COUNTA(%(column_name)s8:%(column_name)s9999)")]
    # data += [formula_row(protocol=protocol, function="=COUNTIF(%(column_name)s8:%(column_name)s9999,1)")]
    # data += [formula_row(protocol=protocol, function="=COUNTIF(%(column_name)s8:%(column_name)s9999,0)")]
    # data += [formula_row(protocol=protocol, function="=COUNTIF(%(column_name)s8:%(column_name)s9999,\"?\")")]
    # add an empty row for clarity
    data += [[]]
    data += [category_headers(protocol)]
    data += [headers(protocol)]
    data += urllistreport_to_spreadsheet_data(category_name=report.urllist.name, urls=urls, protocol=protocol)

    filename = "internet nl dashboard report " \
               f"{report.pk} {report.urllist.name} {report.urllist.scan_type} {report.at_when.date()}"

    # The sheet is created into memory and then passed to the caller. They may save it, or serve it, etc...
    # http://docs.pyexcel.org/en/latest/tutorial06.html?highlight=memory
    # An issue with Sheet prevents uneven row lengths to be added. But that works fine with bookdicts
    # The name of a sheet can only be 32 characters, make sure the most significant info is in the title first.
    tabname = f"{report.pk} {report.urllist.scan_type} {report.at_when.date()} {report.urllist.name}"[0:31]
    book = p.get_book(bookdict={slugify(tabname): data})

    return filename, book


def upgrade_excel_spreadsheet(spreadsheet_data):

    with NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        log.debug(f"Saving temp outout to {tmp.name}")
        spreadsheet_data.save_as(array=spreadsheet_data, filename=tmp.name)

        workbook = load_workbook(tmp.name)
        worksheet = workbook.active

        # nicer columns
        worksheet.column_dimensions["A"].width = "30"
        worksheet.column_dimensions["B"].width = "30"

        # Add statistic rows:
        worksheet.insert_rows(0, amount=9)

        worksheet['B1'] = "Total"
        worksheet['B2'] = "Passed"
        worksheet['B3'] = "Info"
        worksheet['B4'] = "Warning"
        worksheet['B5'] = "Failed"
        worksheet['B6'] = "Not tested"
        worksheet['B7'] = "Error"
        worksheet['B8'] = "Test not applicable (mail only)"
        worksheet['B9'] = "Percentage passed"

        # bold totals:
        for i in range(1, 10):
            worksheet[f'B{i}'].font = Font(bold=True)

        data_columns = [
            'F', 'G', 'H', 'I', 'J', 'K', 'L', "M", "N", 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z',
            'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK', 'AL', 'AM', 'AN', 'AO',
            'AP', 'AQ', 'AR', 'AS', 'AT', 'AU', 'AV', 'AW', 'AX', 'AY', 'AZ', 'BA', 'BB', 'BC', 'BD',
            'BE', 'BF', 'BG', 'BH', 'BI', 'BJ', 'BK'
        ]

        # add some statistics
        for cell in data_columns:
            # if header, then aggregate
            if worksheet[f'{cell}12'].value:
                # There is a max of 5000 domains per scan. So we set this to something lower.
                # There is no good support of headers versus data, which makes working with excel a drama
                # If you ever read this code, and want a good spreadsheet editor: try Apple Numbers. It's fantastic.
                worksheet[f'{cell}1'] = f'=COUNTA({cell}13:{cell}5050)'
                # todo: also support other values
                worksheet[f'{cell}2'] = f'=COUNTIF({cell}13:{cell}5050, "passed")'
                worksheet[f'{cell}3'] = f'=COUNTIF({cell}13:{cell}5050, "info")'
                worksheet[f'{cell}4'] = f'=COUNTIF({cell}13:{cell}5050, "warning")'
                worksheet[f'{cell}5'] = f'=COUNTIF({cell}13:{cell}5050, "failed")'
                worksheet[f'{cell}6'] = f'=COUNTIF({cell}13:{cell}5050, "not_tested")'
                worksheet[f'{cell}7'] = f'=' \
                    f'COUNTIF({cell}13:{cell}5050, "error")+' \
                    f'COUNTIF({cell}13:{cell}5050, "unreachable")+' \
                    f'COUNTIF({cell}13:{cell}5050, "untestable")+' \
                    f'COUNTIF({cell}13:{cell}5050, "not_testable")'
                worksheet[f'{cell}8'] = f'=' \
                    f'COUNTIF({cell}13:{cell}5050, "no_mx")+' \
                    f'COUNTIF({cell}13:{cell}5050, "not_applicable")'
                # Not applicable and not testable are subtracted from the total.
                # See https://github.com/internetstandards/Internet.nl-dashboard/issues/68
                # Rounding's num digits is NOT the number of digits behind the comma, but the total number of digits.
                # todo: we should use the calculations in report.py. And there include the "missing" / empty stuff IF
                # that is missing.
                #                   IF(     H1=0,0,ROUND(     H2÷     H1, 4))
                worksheet[f'{cell}9'] = f'=IF({cell}1=0,0,ROUND({cell}2/{cell}1, 4))'
                worksheet[f'{cell}9'].number_format = '0.00%'

        # make headers bold
        worksheet['A12'].font = Font(bold=True)  # List
        worksheet['B12'].font = Font(bold=True)  # Url
        worksheet['C11'].font = Font(bold=True)  # overall
        worksheet['C12'].font = Font(bold=True)  # % Score
        worksheet['D12'].font = Font(bold=True)  # Report
        for cell in data_columns:
            worksheet[f'{cell}11'].font = Font(bold=True)
            worksheet[f'{cell}12'].font = Font(bold=True)

        # Freeze pane to make navigation easier.
        worksheet.freeze_panes = worksheet['E13']

        # there is probably a feature that puts this in a single conditional value.
        conditional_rules = {
            "passed": PatternFill(start_color='B7FFC8', end_color='B7FFC8', fill_type='solid'),
            "failed": PatternFill(start_color='FFB7B7', end_color='FFB7B7', fill_type='solid'),
            "warning": PatternFill(start_color='FFD9B7', end_color='FFD9B7', fill_type='solid'),
            "info": PatternFill(start_color='B7E3FF', end_color='B7E3FF', fill_type='solid'),
            "good_not_tested": PatternFill(start_color='99FFFF', end_color='C0C0C0', fill_type='solid'),
            "not_tested": PatternFill(start_color='99FFFF', end_color='DBDBDB', fill_type='solid'),
        }

        # Set the measurements to green/red depending on value using conditional formatting.
        # There is no true/false, but we can color based on value.
        for grade, pattern in conditional_rules.items():
            worksheet.conditional_formatting.add(
                'F13:CD5050',
                CellIsRule(operator='=', formula=[f'"{grade}"'], stopIfTrue=True, fill=pattern)
            )

        workbook.save(tmp.name)

        return tmp


def category_headers(protocol: str = 'dns_soa'):
    sheet_headers: List[str] = ['', '']
    for group in SANE_COLUMN_ORDER[protocol]:
        sheet_headers += [translate_field(group, translation_dictionary=po_file_as_dictionary)]

        for _ in range(len(SANE_COLUMN_ORDER[protocol][group])-1):
            sheet_headers += ['']

        # add empty thing after each group to make distinction per group clearer
        sheet_headers += ['']

    return sheet_headers


def headers(protocol: str = 'dns_soa'):
    sheet_headers = ['List', 'Url']
    for group in SANE_COLUMN_ORDER[protocol]:
        sheet_headers += SANE_COLUMN_ORDER[protocol][group]
        # add empty thing after each group to make distinction per group clearer
        sheet_headers += ['']

    # translate them:
    sheet_headers = [translate_field(header, translation_dictionary=po_file_as_dictionary) for header in sheet_headers]

    return sheet_headers


def formula_row(function: str, protocol: str = 'dns_soa'):
    data = []

    my_headers = headers(protocol)
    total = len(my_headers)

    empty_headers = []
    for i in range(total):
        if my_headers[i] == '':
            empty_headers.append(i)

    # log.debug(empty_headers)

    # there is probably a function that does both. I'm not very familiar with itertools.
    index = 0
    for column_name in itertools.islice(iter_all_strings(), total):
        if index in empty_headers:
            data.append('')
        else:
            data.append(function % {'column_name': column_name})

        index += 1

    return data


def urllistreport_to_spreadsheet_data(category_name: str, urls: List[Any], protocol: str = 'dns_soa'):
    data = []
    for url in urls:

        if len(url['endpoints']) == 1:
            # we can just put the whole result in one, which is nicer to look at.
            for endpoint in url['endpoints']:
                if endpoint['protocol'] != protocol:
                    continue
                keyed_ratings = endpoint['ratings_by_type']
                data.append([category_name, url['url']] + keyed_values_as_boolean(keyed_ratings, protocol))

        else:
            data.append([category_name, url['url']])

            for endpoint in url['endpoints']:
                if endpoint['protocol'] != protocol:
                    continue
                keyed_ratings = endpoint['ratings_by_type']
                data.append(['', ''] + keyed_values_as_boolean(keyed_ratings, protocol))

    # log.debug(data)
    return data


def keyed_values_as_boolean(keyed_ratings: Dict[str, Dict[str, Union[str, int]]], protocol: str = 'dns_soa'):
    """
    Keyed rating:
    {'internet_nl_mail_auth_dkim_exist': {'comply_or_explain_explained_on': '',
                                      'comply_or_explain_explanation': '',
                                      'comply_or_explain_explanation_valid_until': '',
                                      'comply_or_explain_valid_at_time_of_report': False,
                                      'explanation': 'Test '
                                                     'internet_nl_mail_auth_dkim_exist '
                                                     'resulted in failed.',
                                      'high': 1,
                                      'is_explained': False,
                                      'last_scan': '2019-07-09T11:07:43.510452+00:00',
                                      'low': 0,
                                      'medium': 0,
                                      'not_applicable': False,
                                      'not_testable': False,
                                      'ok': 0,
                                      'scan': 43945,
                                      'scan_type': 'internet_nl_mail_auth_dkim_exist',
                                      'since': '2019-07-09T11:07:43.510175+00:00',
                                      'type': 'internet_nl_mail_auth_dkim_exist'},...

    :param keyed_ratings:
    :param protocol:
    :return:
    """

    values = []

    for group in SANE_COLUMN_ORDER[protocol]:
        for issue_name in SANE_COLUMN_ORDER[protocol][group]:
            values.append(some_value(issue_name, keyed_ratings))

            if issue_name == 'internet_nl_score_report':
                # add empty column
                values.append(" ")

        # add empty thing after each group to make distinction per group clearer
        # overall group already adds an extra value (url), so we don't need this.
        if group != "overall":
            values += ['']

    return values


def some_value(issue_name: str, keyed_ratings: Dict[str, Dict[str, Union[str, int]]]) -> Union[str, int]:

    if issue_name == 'internet_nl_score':
        # Handle the special case of the score column.
        # explanation":"75 https://batch.internet.nl/mail/portaal.digimelding.nl/289480/",
        # Not steadily convertable to a percentage, so printing it as an integer instead.
        score = keyed_ratings[issue_name]['internet_nl_score']
        return score if score == "error" else int(score)

    if issue_name == 'internet_nl_score_report':
        # fake column to give the column a title per #205, also makes the report more explicit.
        return keyed_ratings['internet_nl_score']['internet_nl_url']

    # the issue name might not exist, the 'ok' value might not exist. In those cases replace it with a ?
    value = keyed_ratings.get(issue_name, None)
    if not value:
        return "?"

    # api v2, tls1.3 update
    if value.get('test_result', False):
        test_value = value.get('test_result', '?')
        # per 205, translate not_testable to untestable. This is cosmetic as the 'not_testable' is
        # everywhere in the software and is just renamed, and will probably be renamed a few times
        # more in the future.
        if test_value == "not_testable":
            test_value = "untestable"
        return test_value

    # unknown columns and data will be empty.
    if "simple_verdict" not in value:
        return ''

    # backward compatible with api v1 reports unreachable
    mapping: Dict[str, str] = {
        'not_testable': 'untestable',
        'not_applicable': 'not_applicable',
        'error_in_test': 'error'
    }
    return mapping.get(str(value['simple_verdict']), "?")
